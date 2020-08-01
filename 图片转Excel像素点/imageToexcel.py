from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill, Fill

# 图片最好不要超640*480!!!!!!!!!!
imageFileName = 'ycy.jpg'  # 图片文件名
image = Image.open(imageFileName)  # 打开图片
wb = openpyxl.Workbook()  # 创建Excel
sheet = wb.create_sheet(imageFileName)  # 创建sheet
sheet.column_dimensions.width = 20.0
sheet.row_dimensions.height = 20.0
imgW, imgH = image.size  # 获取图片大小
print(imgW)
for w in range(imgW):
    print(w)
    print("---------------")
    print(imgW)
    for h in range(imgH):
        # 将每个像素的颜色填充到对应cell的背景色中
        rgba = image.getpixel((w, h))
        colorHex = hex(rgba[0])[2:].zfill(2) + hex(rgba[1])[2:].zfill(2) + hex(rgba[2])[2:].zfill(2)
        fill = PatternFill(fill_type='solid', start_color=colorHex, end_color=colorHex)
        sheet.cell(row=h + 1, column=w + 1).fill = fill

wb.save(imageFileName + '.xlsx')  # 保存xlsx文件
print("done!")
# xcel里显示如果有图像变形，是因为电脑分辨率的原因
